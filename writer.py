"""
XLSX and CSV writing with proper formatting and idempotent updates.
"""

import csv
import os
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.styles import numbers, Font, Alignment, Border, Side


COLUMN_HEADERS = [
    'Date Due',
    'Patient Name',
    'Total Units',
    'Unit Price',
    'Alloys/Extras Cost',
    'Total Cost',
]


def get_month_folder(date_due: datetime, base_path: str = None) -> Path:
    """
    Get the folder path for a given month.
    
    Args:
        date_due: Date to determine month
        base_path: Base output path (default ~/Documents/AccudentLab/)
        
    Returns:
        Path to month folder
    """
    if base_path is None:
        base_path = os.path.expanduser('~/Documents/AccudentLab')
    
    base = Path(base_path)
    month_folder = base / date_due.strftime('%Y-%m')
    month_folder.mkdir(parents=True, exist_ok=True)
    
    # Create logs subfolder
    (month_folder / 'logs').mkdir(exist_ok=True)
    
    return month_folder


def write_xlsx(month_folder: Path, rows: List[Dict]) -> Path:
    """
    Write or update XLSX file with invoice data.
    
    Idempotent: Updates existing rows matching (date_due, patient_name, total_cost).
    
    Args:
        month_folder: Path to month folder
        rows: List of parsed invoice dicts
        
    Returns:
        Path to XLSX file
    """
    xlsx_path = month_folder / f"{month_folder.name}_Accudent.xlsx"
    
    # Load existing workbook or create new
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Accudent'
        # Write headers
        for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    
    # Build lookup of existing rows: (date_due, patient_name, total_cost) -> row_idx
    existing_rows = {}
    for row_idx in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=1).value
        name_val = ws.cell(row=row_idx, column=2).value
        total_val = ws.cell(row=row_idx, column=6).value
        
        if date_val and name_val and total_val:
            # Convert date to string for comparison
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)
            
            # Convert total to Decimal
            total_dec = Decimal(str(total_val))
            
            key = (date_str, name_val, total_dec)
            existing_rows[key] = row_idx
    
    # Process each row: update if exists, append if new
    for row_data in rows:
        date_str = row_data['date_due'].strftime('%Y-%m-%d')
        key = (date_str, row_data['patient_name'], row_data['total_cost'])
        
        if key in existing_rows:
            # Update existing row
            row_idx = existing_rows[key]
        else:
            # Append new row
            row_idx = ws.max_row + 1
        
        _write_row(ws, row_idx, row_data)
    
    # Sort by Date Due (column A)
    _sort_by_date(ws)
    
    # Apply formatting
    _apply_formats(ws)
    
    # Save
    wb.save(xlsx_path)
    
    return xlsx_path


def _write_row(ws, row_idx: int, row_data: Dict):
    """Write a single data row to worksheet."""
    # A: Date Due (datetime object, will be formatted)
    ws.cell(row=row_idx, column=1, value=row_data['date_due'])
    
    # B: Patient Name
    ws.cell(row=row_idx, column=2, value=row_data['patient_name'])
    
    # C: Total Units (integer)
    ws.cell(row=row_idx, column=3, value=row_data['total_units'])
    
    # D: Unit Price (Decimal or None)
    if row_data['unit_price'] is not None:
        ws.cell(row=row_idx, column=4, value=float(row_data['unit_price']))
    else:
        ws.cell(row=row_idx, column=4, value=None)
    
    # E: Alloys/Extras Cost (Decimal)
    ws.cell(row=row_idx, column=5, value=float(row_data['alloys_extras_cost']))
    
    # F: Total Cost (Decimal)
    ws.cell(row=row_idx, column=6, value=float(row_data['total_cost']))
    
    # G: Hidden helper column (formula for legacy parity)
    # Formula: =IF(C*D(D+E)=0,"",C*(D+E))
    # But we want: =IF(CROW*(DROW+EROW)=0,"",CROW*(DROW+EROW))
    formula = f'=IF(C{row_idx}*(D{row_idx}+E{row_idx})=0,"",C{row_idx}*(D{row_idx}+E{row_idx}))'
    ws.cell(row=row_idx, column=7, value=formula)


def _sort_by_date(ws):
    """Sort worksheet by Date Due (column A) ascending."""
    # Get all data rows (skip header)
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Only include rows with data
        if row[0].value:  # Has date
            data_rows.append([cell.value for cell in row])
    
    if not data_rows:
        return
    
    # Sort by first column (date)
    data_rows.sort(key=lambda r: r[0] if isinstance(r[0], datetime) else datetime.max)
    
    # Write back sorted data
    for idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col_idx)
            # Re-apply formula for column G
            if col_idx == 7 and isinstance(value, str) and value.startswith('='):
                cell.value = value
            else:
                cell.value = value


def _apply_formats(ws):
    """Apply number formats and styling to columns."""
    # Define styles
    header_font = Font(name='Helvetica', size=11, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Helvetica', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style header row
    for col_idx in range(1, 7):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Format and style data rows
    for row_idx in range(2, ws.max_row + 1):
        # A: Date as m/d/yyyy
        cell_a = ws.cell(row=row_idx, column=1)
        cell_a.number_format = 'M/D/YYYY'
        cell_a.alignment = center_alignment
        cell_a.font = data_font
        cell_a.border = thin_border
        
        # B: Patient Name
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.alignment = left_alignment
        cell_b.font = data_font
        cell_b.border = thin_border
        
        # C: Total Units as integer
        cell_c = ws.cell(row=row_idx, column=3)
        cell_c.number_format = '0'
        cell_c.alignment = right_alignment
        cell_c.font = data_font
        cell_c.border = thin_border
        
        # D: Unit Price as currency (if present)
        cell_d = ws.cell(row=row_idx, column=4)
        if cell_d.value is not None:
            cell_d.number_format = '$#,##0.00'
        cell_d.alignment = right_alignment
        cell_d.font = data_font
        cell_d.border = thin_border
        
        # E: Alloys/Extras Cost as currency
        cell_e = ws.cell(row=row_idx, column=5)
        cell_e.number_format = '$#,##0.00'
        cell_e.alignment = right_alignment
        cell_e.font = data_font
        cell_e.border = thin_border
        
        # F: Total Cost as currency
        cell_f = ws.cell(row=row_idx, column=6)
        cell_f.number_format = '$#,##0.00'
        cell_f.alignment = right_alignment
        cell_f.font = data_font
        cell_f.border = thin_border
        
        # G: Hidden helper (currency)
        cell_g = ws.cell(row=row_idx, column=7)
        cell_g.number_format = '$#,##0.00'
    
    # Set column widths for better readability
    ws.column_dimensions['A'].width = 12  # Date Due
    ws.column_dimensions['B'].width = 25  # Patient Name
    ws.column_dimensions['C'].width = 12  # Total Units
    ws.column_dimensions['D'].width = 13  # Unit Price
    ws.column_dimensions['E'].width = 18  # Alloys/Extras Cost
    ws.column_dimensions['F'].width = 13  # Total Cost
    
    # Hide column G
    ws.column_dimensions['G'].hidden = True
    
    # Freeze header row for scrolling
    ws.freeze_panes = 'A2'
    
    # Enable auto-filter on header row
    ws.auto_filter.ref = f'A1:F{ws.max_row}'


def write_csv(month_folder: Path, rows: List[Dict]) -> Path:
    """
    Write CSV mirror of XLSX (columns A-F only, no hidden column G).
    
    Args:
        month_folder: Path to month folder
        rows: List of parsed invoice dicts (should be sorted)
        
    Returns:
        Path to CSV file
    """
    csv_path = month_folder / f"{month_folder.name}_Accudent.csv"
    
    # Sort rows by date
    sorted_rows = sorted(rows, key=lambda r: r['date_due'])
    
    with open(csv_path, 'w', newline='') as f:
        writer = csv.writer(f)
        
        # Write header
        writer.writerow(COLUMN_HEADERS)
        
        # Write data rows
        for row_data in sorted_rows:
            # Format date as m/d/yyyy
            date_str = row_data['date_due'].strftime('%-m/%-d/%Y')
            
            # Format unit price (blank if None)
            unit_price_str = f"${row_data['unit_price']:.2f}" if row_data['unit_price'] else ''
            
            writer.writerow([
                date_str,
                row_data['patient_name'],
                row_data['total_units'],
                unit_price_str,
                f"${row_data['alloys_extras_cost']:.2f}",
                f"${row_data['total_cost']:.2f}",
            ])
    
    return csv_path


def load_existing_rows(month_folder: Path) -> List[Dict]:
    """
    Load existing rows from XLSX if it exists.
    
    Args:
        month_folder: Path to month folder
        
    Returns:
        List of row dicts
    """
    xlsx_path = month_folder / f"{month_folder.name}_Accudent.xlsx"
    
    if not xlsx_path.exists():
        return []
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=1).value
        if not date_val:
            continue
        
        # Parse row
        unit_price_val = ws.cell(row=row_idx, column=4).value
        
        rows.append({
            'dentist_name': '',  # Not stored in Excel, only used for PDF report
            'date_due': date_val if isinstance(date_val, datetime) else datetime.strptime(str(date_val), '%Y-%m-%d'),
            'patient_name': ws.cell(row=row_idx, column=2).value,
            'total_units': int(ws.cell(row=row_idx, column=3).value or 0),
            'unit_price': Decimal(str(unit_price_val)) if unit_price_val else None,
            'alloys_extras_cost': Decimal(str(ws.cell(row=row_idx, column=5).value or 0)),
            'total_cost': Decimal(str(ws.cell(row=row_idx, column=6).value or 0)),
        })
    
    return rows
